import openpyxl
def readdata_excel():
    # global username, password
    path = "/Users/rtanne/PycharmProjects/PythonProject_demo/pytest_demo/Login_credentials.xlsx"
    book = openpyxl.load_workbook(path)
    sheet = book.active
    cell_row = sheet.max_row
    cell_col = sheet.max_column
    print(cell_row)
    print(cell_col)
    login_data=[]
    for i in range(2, cell_row+1):
        #for j in range(1, cell_col+1):
            username = sheet.cell(row=i,column=1).value
            password = sheet.cell(row=i, column=2).value
            print(username)
            print(password)
            if username is None or password is None:
                continue
            login_data.append((username,password))
    return login_data
